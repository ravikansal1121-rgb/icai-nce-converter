"""
Parse T-format financial statements from Excel files.
Handles Trading & P/L Account and Balance Sheet in side-by-side (Dr/Cr) layout.
"""
import re
import openpyxl


def _clean_amount(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s in ('-', '—', '–', 'nil', 'Nil', 'NIL', '0'):
        return 0.0
    neg = False
    if s.startswith('(') and s.endswith(')'):
        neg = True
        s = s[1:-1]
    s = s.replace(',', '').replace('₹', '').replace('Rs.', '').replace('Rs', '').strip()
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return 0.0


def _clean_name(val):
    if val is None:
        return ''
    s = str(val).strip()
    s = re.sub(r'^(To|By)\s+', '', s, flags=re.IGNORECASE).strip()
    s = re.sub(r'\s+', ' ', s)
    return s


def _is_section_header(text, keywords):
    if not text:
        return False
    t = text.lower().strip()
    for kw in keywords:
        if kw in t:
            return True
    return False


def _find_amount_col(ws, row, start_col, end_col):
    for c in range(start_col, end_col + 1):
        val = ws.cell(row, c).value
        if isinstance(val, (int, float)) and val != 0:
            return c
        if val and re.match(r'^[\d,]+\.?\d*$', str(val).replace(',', '').strip()):
            return c
    return None


def _detect_layout(ws):
    max_col = min(ws.max_column or 10, 20)
    max_row = min(ws.max_row or 50, 100)

    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is not None:
                row_vals.append((c, str(v).strip()))

        texts = [t.lower() for _, t in row_vals]
        has_dr = any('dr' in t or 'debit' in t for t in texts)
        has_cr = any('cr' in t or 'credit' in t for t in texts)
        has_liab = any('liabilit' in t for t in texts)
        has_asset = any('asset' in t for t in texts)

        if has_dr and has_cr:
            cols = [c for c, _ in row_vals]
            if len(cols) >= 2:
                mid = (min(cols) + max(cols)) // 2
                return {
                    'type': 'side_by_side',
                    'header_row': r,
                    'left_name_col': min(cols),
                    'mid_col': mid,
                    'right_max_col': max(cols),
                }
        if has_liab and has_asset:
            cols = [c for c, _ in row_vals]
            if len(cols) >= 2:
                mid = (min(cols) + max(cols)) // 2
                return {
                    'type': 'side_by_side',
                    'header_row': r,
                    'left_name_col': min(cols),
                    'mid_col': mid,
                    'right_max_col': max(cols),
                }
    return {'type': 'vertical', 'header_row': 1}


def _extract_items_from_range(ws, start_row, end_row, name_col, amt_col_start, amt_col_end):
    items = []
    for r in range(start_row, end_row + 1):
        name_val = ws.cell(r, name_col).value
        name = _clean_name(name_val)
        if not name:
            continue

        skip_patterns = [
            r'^total', r'^grand total', r'^net profit', r'^net loss',
            r'^gross profit', r'^gross loss', r'^balance c/d',
            r'^balance b/d', r'^as per', r'^amount',
            r'^particulars', r'^dr\.?$', r'^cr\.?$',
            r'^assets?$', r'^liabilit', r'^equity',
            r'^current assets?$', r'^non.?current', r'^fixed assets?$',
        ]
        if any(re.match(p, name.lower()) for p in skip_patterns):
            amt = 0.0
            for c in range(amt_col_start, amt_col_end + 1):
                a = _clean_amount(ws.cell(r, c).value)
                if a != 0:
                    amt = a
                    break
            if 'gross profit' in name.lower() or 'net profit' in name.lower():
                items.append({'name': name, 'amount': abs(amt), 'is_total': True})
            continue

        amt = 0.0
        for c in range(amt_col_start, amt_col_end + 1):
            a = _clean_amount(ws.cell(r, c).value)
            if a != 0:
                amt = a
                break

        if amt != 0 or (name and len(name) > 2):
            items.append({'name': name, 'amount': abs(amt), 'is_total': False})

    return [i for i in items if i['amount'] != 0 or not i.get('is_total')]


def _find_sections(ws):
    sections = {}
    max_row = ws.max_row or 1
    max_col = min(ws.max_column or 6, 20)

    trading_kw = ['trading', 'trading account', 'trading and profit', 'trading & profit']
    pl_kw = ['profit and loss', 'profit & loss', 'income and expenditure']
    bs_kw = ['balance sheet', 'balancesheet']
    capital_kw = ['capital account', 'capital a/c']

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = ws.cell(r, c).value
            if not val:
                continue
            t = str(val).strip().lower()

            if _is_section_header(t, trading_kw) or _is_section_header(t, pl_kw):
                if 'trading_pl' not in sections:
                    sections['trading_pl'] = {'start': r}
            elif _is_section_header(t, bs_kw):
                sections['balance_sheet'] = {'start': r}
                if 'trading_pl' in sections and 'end' not in sections['trading_pl']:
                    sections['trading_pl']['end'] = r - 1
            elif _is_section_header(t, capital_kw):
                sections['capital'] = {'start': r}
                if 'trading_pl' in sections and 'end' not in sections['trading_pl']:
                    sections['trading_pl']['end'] = r - 1

    if 'trading_pl' in sections and 'end' not in sections['trading_pl']:
        if 'balance_sheet' in sections:
            sections['trading_pl']['end'] = sections['balance_sheet']['start'] - 1
        elif 'capital' in sections:
            sections['trading_pl']['end'] = sections['capital']['start'] - 1
        else:
            sections['trading_pl']['end'] = max_row

    if 'balance_sheet' in sections and 'end' not in sections['balance_sheet']:
        if 'capital' in sections:
            sections['balance_sheet']['end'] = sections['capital']['start'] - 1
        else:
            sections['balance_sheet']['end'] = max_row

    if 'capital' in sections and 'end' not in sections['capital']:
        sections['capital']['end'] = max_row

    return sections


def _parse_t_format_section(ws, start_row, end_row, max_col):
    mid_col = max_col // 2 if max_col > 3 else 3

    left_name_col = 1
    left_amt_end = mid_col
    right_name_col = mid_col + 1
    right_amt_end = max_col

    left_headers = ('dr', 'dr.', 'debit', 'debit side', 'liabilities', 'liability')
    right_headers = ('cr', 'cr.', 'credit', 'credit side', 'assets', 'asset')

    for r in range(start_row, min(start_row + 5, end_row + 1)):
        for c in range(1, max_col + 1):
            val = ws.cell(r, c).value
            if val:
                t = str(val).strip().lower()
                if t in left_headers:
                    left_name_col = c
                elif t in right_headers:
                    right_name_col = c
                    left_amt_end = c - 1
                    right_amt_end = max_col

    if right_name_col <= left_name_col + 1:
        for r in range(start_row + 1, min(start_row + 10, end_row + 1)):
            found_left = False
            found_right_col = 0
            for c in range(1, max_col + 1):
                val = ws.cell(r, c).value
                if val and not isinstance(val, (int, float)):
                    if not found_left:
                        found_left = True
                    elif c > left_name_col + 1:
                        found_right_col = c
                        break
            if found_right_col > 0:
                right_name_col = found_right_col
                left_amt_end = found_right_col - 1
                right_amt_end = max_col
                break

    for c in range(left_name_col + 1, right_name_col):
        for r in range(start_row + 1, min(start_row + 15, end_row + 1)):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)) and v > 0:
                left_amt_end = max(left_amt_end, c)
                break

    debit_items = _extract_items_from_range(
        ws, start_row + 1, end_row, left_name_col, left_name_col + 1, left_amt_end
    )
    credit_items = _extract_items_from_range(
        ws, start_row + 1, end_row, right_name_col, right_name_col + 1, right_amt_end
    )

    return debit_items, credit_items


def _parse_vertical_section(ws, start_row, end_row, max_col):
    items = []
    for r in range(start_row + 1, end_row + 1):
        name_val = ws.cell(r, 1).value
        if not name_val:
            name_val = ws.cell(r, 2).value
        name = _clean_name(name_val)
        if not name:
            continue

        amt = 0.0
        for c in range(2, max_col + 1):
            a = _clean_amount(ws.cell(r, c).value)
            if a != 0:
                amt = a
                break

        if amt != 0:
            items.append({'name': name, 'amount': amt})

    debit = [i for i in items if i['amount'] > 0]
    credit = [i for i in items if i['amount'] < 0]
    for c in credit:
        c['amount'] = abs(c['amount'])
    return debit, credit


def _extract_entity_name(ws):
    for r in range(1, 6):
        for c in range(1, 10):
            val = ws.cell(r, c).value
            if val:
                t = str(val).strip()
                t_lower = t.lower()
                skip = ['trading', 'profit', 'balance', 'capital',
                        'for the', 'as at', 'year ended', 'dr', 'cr',
                        'particulars', 'amount', 'schedule']
                if not any(s in t_lower for s in skip) and len(t) > 3:
                    if not re.match(r'^[\d,.\s]+$', t):
                        return t
    return 'Entity Name'


def _extract_year(ws):
    for r in range(1, 10):
        for c in range(1, 15):
            val = ws.cell(r, c).value
            if val:
                t = str(val).strip()
                m = re.search(r'(\d{4})\s*[-–]\s*(\d{2,4})', t)
                if m:
                    y1 = m.group(1)
                    y2 = m.group(2)
                    if len(y2) == 2:
                        y2 = y1[:2] + y2
                    return f'{y1}-{y2[-2:]}'
                m = re.search(r'31\s*(st)?\s*march\s*,?\s*(\d{4})', t, re.IGNORECASE)
                if m:
                    y = int(m.group(2))
                    return f'{y-1}-{str(y)[-2:]}'
                m = re.search(r'(\d{4})', t)
                if m and 'march' in t.lower():
                    y = int(m.group(1))
                    return f'{y-1}-{str(y)[-2:]}'
    return ''


def _infer_constitution(data):
    all_names = []
    for section in ['trading_pl', 'balance_sheet', 'capital']:
        if section in data:
            sd = data[section]
            if isinstance(sd, dict):
                for side in ['debit', 'credit', 'liabilities', 'assets']:
                    if side in sd:
                        all_names.extend([i['name'].lower() for i in sd[side]])

    entity = data.get('entity_name', '').lower()
    if '& co' in entity or 'and co' in entity:
        return 'partnership'

    for n in all_names:
        if 'partner' in n:
            return 'partnership'

    return 'proprietorship'


def _infer_proprietor(data):
    entity = data.get('entity_name', '')
    if data.get('constitution') == 'partnership':
        return ''
    words = ['enterprise', 'enterprises', 'traders', 'trading',
             'agencies', 'store', 'stores', 'mart', 'emporium',
             'industries', 'works', 'co', 'company']
    name = entity
    for w in words:
        name = re.sub(r'\b' + w + r'\b', '', name, flags=re.IGNORECASE).strip()
    name = re.sub(r'[&,.\-]+$', '', name).strip()
    name = re.sub(r'^(M/s\.?\s*|Shri\s+|Smt\.?\s+)', '', name).strip()
    return name if len(name) > 2 else entity


def _parse_capital_schedule(ws, start_row, end_row, max_col):
    cap = {
        'opening': 0, 'capital_introduced': 0, 'net_profit': 0,
        'interest_on_capital': 0, 'drawings': 0, 'closing': 0,
    }
    for r in range(start_row, end_row + 1):
        name = ''
        amt = 0.0
        for c in range(1, max_col + 1):
            val = ws.cell(r, c).value
            if val and not isinstance(val, (int, float)):
                name = str(val).strip().lower()
            if isinstance(val, (int, float)) and val != 0:
                amt = float(val)

        if not name:
            continue
        if 'opening' in name or 'balance b/d' in name:
            cap['opening'] = abs(amt)
        elif 'capital introduced' in name or 'additional capital' in name or 'brought in' in name:
            cap['capital_introduced'] = abs(amt)
        elif 'net profit' in name or 'profit for' in name:
            cap['net_profit'] = abs(amt)
        elif 'interest on capital' in name:
            cap['interest_on_capital'] = abs(amt)
        elif 'drawing' in name or 'withdrawal' in name:
            cap['drawings'] = abs(amt)
        elif 'closing' in name or 'balance c/d' in name:
            cap['closing'] = abs(amt)

    if cap['closing'] == 0 and cap['opening'] > 0:
        cap['closing'] = (cap['opening'] + cap['capital_introduced'] +
                          cap['net_profit'] + cap['interest_on_capital'] - cap['drawings'])

    return cap


def parse_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    results = []

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        if ws.max_row is None or ws.max_row < 3:
            continue

        entity_name = _extract_entity_name(ws)
        year = _extract_year(ws)
        sections = _find_sections(ws)
        max_col = min(ws.max_column or 6, 20)

        data = {
            'entity_name': entity_name,
            'fy': year,
            'source_sheet': ws_name,
        }

        if 'trading_pl' in sections:
            sec = sections['trading_pl']
            debit, credit = _parse_t_format_section(ws, sec['start'], sec['end'], max_col)
            if not debit and not credit:
                debit, credit = _parse_vertical_section(ws, sec['start'], sec['end'], max_col)
            data['trading_pl'] = {'debit': debit, 'credit': credit}

        if 'balance_sheet' in sections:
            sec = sections['balance_sheet']
            liab, assets = _parse_t_format_section(ws, sec['start'], sec['end'], max_col)
            if not liab and not assets:
                liab, assets = _parse_vertical_section(ws, sec['start'], sec['end'], max_col)
            data['balance_sheet'] = {'liabilities': liab, 'assets': assets}

        if 'capital' in sections:
            sec = sections['capital']
            data['capital_account'] = _parse_capital_schedule(ws, sec['start'], sec['end'], max_col)

        if not sections:
            all_items = []
            for r in range(1, ws.max_row + 1):
                for c in range(1, max_col + 1):
                    val = ws.cell(r, c).value
                    if val and not isinstance(val, (int, float)):
                        name = _clean_name(val)
                        if name and len(name) > 2:
                            amt = 0.0
                            for ac in range(c + 1, max_col + 1):
                                a = _clean_amount(ws.cell(r, ac).value)
                                if a != 0:
                                    amt = a
                                    break
                            if amt != 0:
                                all_items.append({'name': name, 'amount': amt})
            if all_items:
                data['raw_items'] = all_items

        if any(k in data for k in ['trading_pl', 'balance_sheet', 'capital_account', 'raw_items']):
            data['constitution'] = _infer_constitution(data)
            data['proprietor_name'] = _infer_proprietor(data)
            results.append(data)

    wb.close()
    return results
