"""
Build ICAI NCE-format Excel workbook with formula linkage.
Calibri Light 11, no underlines, gridlines off, strict formatting.
"""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

FONT = Font(name='Calibri Light', size=11)
FONT_BOLD = Font(name='Calibri Light', size=11, bold=True)
FONT_ITALIC = Font(name='Calibri Light', size=11, italic=True)
FONT_BOLD_ITALIC = Font(name='Calibri Light', size=11, bold=True, italic=True)

NUM_FMT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

THIN_BOTTOM = Border(bottom=Side(style='thin'))
THIN_TOP_BOTTOM = Border(top=Side(style='thin'), bottom=Side(style='thin'))
DOUBLE_BOTTOM = Border(bottom=Side(style='double'))
THIN_TOP_DOUBLE_BOTTOM = Border(top=Side(style='thin'), bottom=Side(style='double'))


def _apply_font(ws, row, max_col, font=FONT):
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        cell.font = font
        cell.alignment = Alignment(vertical='top', wrap_text=True)


def _set_num(ws, row, col, value=None, formula=None, bold=False):
    cell = ws.cell(row, col)
    if formula:
        cell.value = formula
    elif value is not None:
        cell.value = value
    cell.number_format = NUM_FMT
    cell.font = FONT_BOLD if bold else FONT
    cell.alignment = Alignment(horizontal='right', vertical='top')
    return cell


def _set_text(ws, row, col, text, bold=False, italic=False, indent=0):
    cell = ws.cell(row, col)
    cell.value = text
    if bold and italic:
        cell.font = FONT_BOLD_ITALIC
    elif bold:
        cell.font = FONT_BOLD
    elif italic:
        cell.font = FONT_ITALIC
    else:
        cell.font = FONT
    cell.alignment = Alignment(
        horizontal='left', vertical='top',
        wrap_text=True, indent=indent
    )
    return cell


def _set_center(ws, row, col, text, bold=False):
    cell = ws.cell(row, col)
    cell.value = text
    cell.font = FONT_BOLD if bold else FONT
    cell.alignment = Alignment(horizontal='center', vertical='top')
    return cell


def _apply_border_row(ws, row, max_col, border):
    for c in range(1, max_col + 1):
        ws.cell(row, c).border = border


def _title_block(ws, entity_name, subtitle, max_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    _set_text(ws, 1, 1, entity_name, bold=True)
    ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    _set_text(ws, 2, 1, subtitle, bold=True)
    ws.cell(2, 1).alignment = Alignment(horizontal='center', vertical='center')

    _set_text(ws, 3, max_col, '(Amount in Rs.)', italic=True)
    ws.cell(3, max_col).alignment = Alignment(horizontal='right', vertical='top')

    for r in range(1, 4):
        _apply_font(ws, r, max_col, FONT_BOLD if r <= 2 else FONT_ITALIC)


def _sum_list(items):
    if isinstance(items, list):
        return sum(i.get('amount', 0) for i in items)
    if isinstance(items, (int, float)):
        return items
    return 0


def _setup_sheet(ws):
    ws.sheet_view.showGridLines = False


def build_icai_workbook(classified, note_info, output_path):
    wb = Workbook()

    ws_index = wb.active
    ws_index.title = 'Index'

    ws_bs = wb.create_sheet('Balance Sheet')
    ws_pl = wb.create_sheet('Statement of P&L')

    note_sheets = {}
    retained = note_info['retained']
    note_map = note_info['note_map']

    group_ranges = _group_notes(retained)
    for group_name, notes_in_group in group_ranges:
        ws = wb.create_sheet(group_name)
        _setup_sheet(ws)
        note_sheets[group_name] = ws

    for ws in wb.worksheets:
        _setup_sheet(ws)

    _build_index(ws_index, classified, retained, group_ranges)
    row_refs = {}
    _build_balance_sheet(ws_bs, classified, note_info, row_refs)
    _build_pl(ws_pl, classified, note_info, row_refs)
    _build_notes(note_sheets, classified, note_info, row_refs, group_ranges)

    _normalize_fonts(wb)

    wb.save(output_path)
    return output_path


def _group_notes(retained):
    groups = []
    current_group = []
    current_label_start = None

    category_order = [
        ('entity_policy', ['brief', 'policies', 'capital']),
        ('liability', ['lt_borrowings', 'st_borrowings', 'trade_payables', 'other_cl', 'st_provisions']),
        ('ppe', ['ppe']),
        ('asset', ['non_current_investments', 'lt_loans_advances', 'inventories',
                   'trade_receivables', 'cash_bank', 'st_loans_advances', 'other_ca']),
        ('pl', ['revenue', 'other_income', 'cost_of_materials', 'changes_in_inventories',
                'employee_benefits', 'finance_costs', 'depreciation_note', 'other_expenses_note']),
        ('additional', ['previous_year', 'rounding_off']),
    ]

    retained_ids = {n['id'] for n in retained}

    for cat_name, note_ids in category_order:
        cat_notes = [n for n in retained if n['id'] in note_ids]
        if cat_notes:
            first_num = cat_notes[0]['number']
            last_num = cat_notes[-1]['number']
            if first_num == last_num:
                label = f'Note {first_num}'
            else:
                label = f'Notes {first_num} to {last_num}'
            groups.append((label, cat_notes))

    ungrouped = [n for n in retained if not any(n in g[1] for g in groups)]
    if ungrouped:
        nums = [n['number'] for n in ungrouped]
        label = f'Notes {min(nums)} to {max(nums)}'
        groups.append((label, ungrouped))

    return groups


def _build_index(ws, classified, retained, group_ranges):
    _title_block(ws, classified['entity_name'], 'Index to Financial Statements', 6)

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25

    row = 5
    _set_text(ws, row, 1, 'Sr.', bold=True)
    _set_text(ws, row, 2, 'Particulars', bold=True)
    _set_text(ws, row, 3, 'Sheet Reference', bold=True)
    _apply_border_row(ws, row, 3, THIN_TOP_BOTTOM)
    row += 1

    _set_center(ws, row, 1, 1)
    _set_text(ws, row, 2, 'Balance Sheet')
    _set_text(ws, row, 3, 'Balance Sheet')
    row += 1

    _set_center(ws, row, 1, 2)
    _set_text(ws, row, 2, 'Statement of Profit and Loss')
    _set_text(ws, row, 3, 'Statement of P&L')
    row += 1

    sr = 3
    for group_name, notes in group_ranges:
        _set_center(ws, row, 1, sr)
        note_titles = ', '.join([f"Note {n['number']}: {n['title']}" for n in notes[:3]])
        if len(notes) > 3:
            note_titles += f' ... (+{len(notes)-3} more)'
        _set_text(ws, row, 2, group_name)
        _set_text(ws, row, 3, group_name)
        row += 1
        sr += 1

    _apply_border_row(ws, row - 1, 3, THIN_BOTTOM)


def _build_balance_sheet(ws, classified, note_info, row_refs):
    max_col = 6
    entity = classified['entity_name']
    ye_cy = classified.get('year_ending_cy', '31 March 20XX')
    ye_py = classified.get('year_ending_py', '31 March 20XX')

    _title_block(ws, entity, 'Balance Sheet', max_col)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 2
    ws.column_dimensions['F'].width = 18

    row = 5
    _set_text(ws, row, 2, 'Particulars', bold=True)
    _set_text(ws, row, 3, 'Note', bold=True)
    ws.cell(row, 3).alignment = Alignment(horizontal='center', vertical='top')
    _set_text(ws, row, 4, f'As at {ye_cy}', bold=True)
    ws.cell(row, 4).alignment = Alignment(horizontal='center', vertical='top')
    _set_text(ws, row, 6, f'As at {ye_py}', bold=True)
    ws.cell(row, 6).alignment = Alignment(horizontal='center', vertical='top')
    _apply_border_row(ws, row, max_col, THIN_TOP_BOTTOM)

    row += 1
    bn = note_info['bs_line_to_note']

    row += 1
    _set_text(ws, row, 1, 'I.', bold=True)
    _set_text(ws, row, 2, 'EQUITY AND LIABILITIES', bold=True)
    row += 1

    _set_text(ws, row, 2, "(1) Owner's Capital", bold=True, indent=1)
    row += 1
    cap_note = bn.get('capital', '')
    _set_text(ws, row, 2, 'Capital Account', indent=2)
    _set_center(ws, row, 3, cap_note)
    cap_cy = classified.get('capital_cy', {})
    cap_py = classified.get('capital_py', {})
    cap_closing_cy = cap_cy.get('closing', 0)
    cap_closing_py = cap_py.get('closing', 0)
    _set_num(ws, row, 4, cap_closing_cy)
    _set_num(ws, row, 6, cap_closing_py)
    row_refs['bs_capital_cy'] = (ws.title, row, 4)
    row_refs['bs_capital_py'] = (ws.title, row, 6)

    subtotal_equity_row = row
    row += 1
    _set_text(ws, row, 2, 'Total Equity', bold=True, indent=1)
    _set_num(ws, row, 4, formula=f'=D{subtotal_equity_row}', bold=True)
    _set_num(ws, row, 6, formula=f'=F{subtotal_equity_row}', bold=True)
    _apply_border_row(ws, row, max_col, THIN_BOTTOM)
    total_equity_row_cy = row

    row += 2
    _set_text(ws, row, 2, '(2) Non-Current Liabilities', bold=True, indent=1)
    row += 1
    nc_liab_start = row

    if 'lt_borrowings' in bn:
        _set_text(ws, row, 2, 'Long-Term Borrowings', indent=2)
        _set_center(ws, row, 3, bn['lt_borrowings'])
        _set_num(ws, row, 4, _sum_list(classified.get('bank_loan_cy', [])))
        _set_num(ws, row, 6, _sum_list(classified.get('bank_loan_py', [])))
        row += 1

    nc_liab_end = row - 1
    if nc_liab_end >= nc_liab_start:
        _set_text(ws, row, 2, 'Total Non-Current Liabilities', bold=True, indent=1)
        if nc_liab_end == nc_liab_start:
            _set_num(ws, row, 4, formula=f'=D{nc_liab_start}', bold=True)
            _set_num(ws, row, 6, formula=f'=F{nc_liab_start}', bold=True)
        else:
            _set_num(ws, row, 4, formula=f'=SUM(D{nc_liab_start}:D{nc_liab_end})', bold=True)
            _set_num(ws, row, 6, formula=f'=SUM(F{nc_liab_start}:F{nc_liab_end})', bold=True)
        _apply_border_row(ws, row, max_col, THIN_BOTTOM)
        total_ncl_row = row
        row += 2
    else:
        total_ncl_row = None
        row += 1

    _set_text(ws, row, 2, '(3) Current Liabilities', bold=True, indent=1)
    row += 1
    cl_start = row

    if 'st_borrowings' in bn:
        _set_text(ws, row, 2, 'Short-Term Borrowings', indent=2)
        _set_center(ws, row, 3, bn['st_borrowings'])
        st_borr_cy = _sum_list(classified.get('bank_od_cc_cy', [])) + _sum_list(classified.get('unsecured_loan_cy', []))
        st_borr_py = _sum_list(classified.get('bank_od_cc_py', [])) + _sum_list(classified.get('unsecured_loan_py', []))
        _set_num(ws, row, 4, st_borr_cy)
        _set_num(ws, row, 6, st_borr_py)
        row += 1

    if 'trade_payables' in bn:
        _set_text(ws, row, 2, 'Trade Payables', indent=2)
        _set_center(ws, row, 3, bn['trade_payables'])
        _set_num(ws, row, 4, _sum_list(classified.get('trade_payables_cy', [])))
        _set_num(ws, row, 6, _sum_list(classified.get('trade_payables_py', [])))
        row += 1

    if 'other_cl' in bn:
        _set_text(ws, row, 2, 'Other Current Liabilities', indent=2)
        _set_center(ws, row, 3, bn['other_cl'])
        ocl_cy = (_sum_list(classified.get('gst_payable_cy', [])) +
                  _sum_list(classified.get('tds_payable_cy', [])) +
                  _sum_list(classified.get('other_statutory_cy', [])))
        ocl_py = (_sum_list(classified.get('gst_payable_py', [])) +
                  _sum_list(classified.get('tds_payable_py', [])) +
                  _sum_list(classified.get('other_statutory_py', [])))
        _set_num(ws, row, 4, ocl_cy)
        _set_num(ws, row, 6, ocl_py)
        row += 1

    if 'st_provisions' in bn:
        _set_text(ws, row, 2, 'Short-Term Provisions', indent=2)
        _set_center(ws, row, 3, bn['st_provisions'])
        _set_num(ws, row, 4, _sum_list(classified.get('provision_tax_cy', [])))
        _set_num(ws, row, 6, _sum_list(classified.get('provision_tax_py', [])))
        row += 1

    cl_end = row - 1
    if cl_end >= cl_start:
        _set_text(ws, row, 2, 'Total Current Liabilities', bold=True, indent=1)
        _set_num(ws, row, 4, formula=f'=SUM(D{cl_start}:D{cl_end})', bold=True)
        _set_num(ws, row, 6, formula=f'=SUM(F{cl_start}:F{cl_end})', bold=True)
        _apply_border_row(ws, row, max_col, THIN_BOTTOM)
        total_cl_row = row
    else:
        total_cl_row = None

    row += 2
    _set_text(ws, row, 2, 'TOTAL EQUITY AND LIABILITIES', bold=True)
    parts = [f'D{total_equity_row_cy}']
    if total_ncl_row:
        parts.append(f'D{total_ncl_row}')
    if total_cl_row:
        parts.append(f'D{total_cl_row}')
    _set_num(ws, row, 4, formula='=' + '+'.join(parts), bold=True)

    parts_py = [f'F{total_equity_row_cy}']
    if total_ncl_row:
        parts_py.append(f'F{total_ncl_row}')
    if total_cl_row:
        parts_py.append(f'F{total_cl_row}')
    _set_num(ws, row, 6, formula='=' + '+'.join(parts_py), bold=True)
    _apply_border_row(ws, row, max_col, THIN_TOP_DOUBLE_BOTTOM)
    total_liab_row = row

    row += 3
    _set_text(ws, row, 1, 'II.', bold=True)
    _set_text(ws, row, 2, 'ASSETS', bold=True)
    row += 1

    _set_text(ws, row, 2, '(1) Non-Current Assets', bold=True, indent=1)
    row += 1
    nca_start = row

    if 'ppe' in bn:
        _set_text(ws, row, 2, 'Property, Plant and Equipment', indent=2)
        _set_center(ws, row, 3, bn['ppe'])
        _set_num(ws, row, 4, _sum_list(classified.get('ppe_cy', [])))
        _set_num(ws, row, 6, _sum_list(classified.get('ppe_py', [])))
        row += 1

    if 'non_current_investments' in bn:
        _set_text(ws, row, 2, 'Non-Current Investments', indent=2)
        _set_center(ws, row, 3, bn['non_current_investments'])
        _set_num(ws, row, 4, _sum_list(classified.get('gold_investments_cy', [])))
        _set_num(ws, row, 6, _sum_list(classified.get('gold_investments_py', [])))
        row += 1

    if 'lt_loans_advances' in bn:
        _set_text(ws, row, 2, 'Long-Term Loans and Advances', indent=2)
        _set_center(ws, row, 3, bn['lt_loans_advances'])
        _set_num(ws, row, 4, _sum_list(classified.get('security_deposit_cy', [])))
        _set_num(ws, row, 6, _sum_list(classified.get('security_deposit_py', [])))
        row += 1

    nca_end = row - 1
    if nca_end >= nca_start:
        _set_text(ws, row, 2, 'Total Non-Current Assets', bold=True, indent=1)
        _set_num(ws, row, 4, formula=f'=SUM(D{nca_start}:D{nca_end})', bold=True)
        _set_num(ws, row, 6, formula=f'=SUM(F{nca_start}:F{nca_end})', bold=True)
        _apply_border_row(ws, row, max_col, THIN_BOTTOM)
        total_nca_row = row
        row += 2
    else:
        total_nca_row = None
        row += 1

    _set_text(ws, row, 2, '(2) Current Assets', bold=True, indent=1)
    row += 1
    ca_start = row

    if 'inventories' in bn:
        _set_text(ws, row, 2, 'Inventories', indent=2)
        _set_center(ws, row, 3, bn['inventories'])
        _set_num(ws, row, 4, classified.get('inventories_cy', 0) or classified.get('closing_stock_cy', 0))
        _set_num(ws, row, 6, classified.get('inventories_py', 0) or classified.get('closing_stock_py', 0))
        row += 1

    if 'trade_receivables' in bn:
        _set_text(ws, row, 2, 'Trade Receivables', indent=2)
        _set_center(ws, row, 3, bn['trade_receivables'])
        _set_num(ws, row, 4, _sum_list(classified.get('trade_receivables_cy', [])))
        _set_num(ws, row, 6, _sum_list(classified.get('trade_receivables_py', [])))
        row += 1

    if 'cash_bank' in bn:
        _set_text(ws, row, 2, 'Cash and Bank Balances', indent=2)
        _set_center(ws, row, 3, bn['cash_bank'])
        cb_cy = (_sum_list(classified.get('cash_hand_cy', [])) +
                 _sum_list(classified.get('bank_balance_cy', [])) +
                 _sum_list(classified.get('fixed_deposit_cy', [])))
        cb_py = (_sum_list(classified.get('cash_hand_py', [])) +
                 _sum_list(classified.get('bank_balance_py', [])) +
                 _sum_list(classified.get('fixed_deposit_py', [])))
        _set_num(ws, row, 4, cb_cy)
        _set_num(ws, row, 6, cb_py)
        row += 1

    if 'st_loans_advances' in bn:
        _set_text(ws, row, 2, 'Short-Term Loans and Advances', indent=2)
        _set_center(ws, row, 3, bn['st_loans_advances'])
        stla_cy = _sum_list(classified.get('gst_input_cy', [])) + _sum_list(classified.get('tds_advance_tax_cy', []))
        stla_py = _sum_list(classified.get('gst_input_py', [])) + _sum_list(classified.get('tds_advance_tax_py', []))
        _set_num(ws, row, 4, stla_cy)
        _set_num(ws, row, 6, stla_py)
        row += 1

    if 'other_ca' in bn:
        _set_text(ws, row, 2, 'Other Current Assets', indent=2)
        _set_center(ws, row, 3, bn['other_ca'])
        _set_num(ws, row, 4, _sum_list(classified.get('other_ca_cy', [])))
        _set_num(ws, row, 6, _sum_list(classified.get('other_ca_py', [])))
        row += 1

    ca_end = row - 1
    if ca_end >= ca_start:
        _set_text(ws, row, 2, 'Total Current Assets', bold=True, indent=1)
        _set_num(ws, row, 4, formula=f'=SUM(D{ca_start}:D{ca_end})', bold=True)
        _set_num(ws, row, 6, formula=f'=SUM(F{ca_start}:F{ca_end})', bold=True)
        _apply_border_row(ws, row, max_col, THIN_BOTTOM)
        total_ca_row = row
    else:
        total_ca_row = None

    row += 2
    _set_text(ws, row, 2, 'TOTAL ASSETS', bold=True)
    parts = []
    if total_nca_row:
        parts.append(f'D{total_nca_row}')
    if total_ca_row:
        parts.append(f'D{total_ca_row}')
    _set_num(ws, row, 4, formula='=' + '+'.join(parts) if parts else '=0', bold=True)

    parts_py = []
    if total_nca_row:
        parts_py.append(f'F{total_nca_row}')
    if total_ca_row:
        parts_py.append(f'F{total_ca_row}')
    _set_num(ws, row, 6, formula='=' + '+'.join(parts_py) if parts_py else '=0', bold=True)
    _apply_border_row(ws, row, max_col, THIN_TOP_DOUBLE_BOTTOM)
    total_asset_row = row

    row += 2
    last_note = note_info['last_note_number']
    _set_text(ws, row, 2, f'The accompanying Notes 1 to {last_note} form an integral part of these Financial Statements.',
              italic=True)
    row += 1
    _set_text(ws, row, 2, 'As per our report of even date attached', italic=True)

    row += 2
    _set_text(ws, row, 1, 'For and on behalf of', bold=True)
    _set_text(ws, row, 4, 'For and on behalf of the Entity', bold=True)
    row += 1
    _set_text(ws, row, 1, '____________________')
    _set_text(ws, row, 4, '____________________')
    row += 1
    _set_text(ws, row, 1, 'Chartered Accountants')
    prop_name = classified.get('proprietor_name', '—')
    _set_text(ws, row, 4, f'Shri/Smt. {prop_name}')
    row += 1
    _set_text(ws, row, 1, 'Firm Registration No.: —')
    _set_text(ws, row, 4, '(Proprietor)')
    row += 2
    _set_text(ws, row, 1, 'Place: —')
    _set_text(ws, row, 4, 'Place: —')
    row += 1
    _set_text(ws, row, 1, 'Date: —')
    _set_text(ws, row, 4, 'Date: —')

    row_refs['total_liab_row'] = total_liab_row
    row_refs['total_asset_row'] = total_asset_row


def _build_pl(ws, classified, note_info, row_refs):
    max_col = 6
    entity = classified['entity_name']
    ye_cy = classified.get('year_ending_cy', '31 March 20XX')
    ye_py = classified.get('year_ending_py', '31 March 20XX')

    _title_block(ws, entity, 'Statement of Profit and Loss', max_col)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 2
    ws.column_dimensions['F'].width = 18

    row = 5
    _set_text(ws, row, 2, 'Particulars', bold=True)
    _set_text(ws, row, 3, 'Note', bold=True)
    ws.cell(row, 3).alignment = Alignment(horizontal='center', vertical='top')
    _set_text(ws, row, 4, f'For the year ended {ye_cy}', bold=True)
    ws.cell(row, 4).alignment = Alignment(horizontal='center', vertical='top')
    _set_text(ws, row, 6, f'For the year ended {ye_py}', bold=True)
    ws.cell(row, 6).alignment = Alignment(horizontal='center', vertical='top')
    _apply_border_row(ws, row, max_col, THIN_TOP_BOTTOM)

    pn = note_info['pl_line_to_note']
    row += 2

    _set_text(ws, row, 1, 'I.', bold=True)
    _set_text(ws, row, 2, 'Revenue from Operations', indent=1)
    if 'revenue' in pn:
        _set_center(ws, row, 3, pn['revenue'])
    _set_num(ws, row, 4, _sum_list(classified.get('revenue_from_operations_cy', [])))
    _set_num(ws, row, 6, _sum_list(classified.get('revenue_from_operations_py', [])))
    rev_row = row
    row += 1

    _set_text(ws, row, 1, 'II.', bold=True)
    _set_text(ws, row, 2, 'Other Income', indent=1)
    if 'other_income' in pn:
        _set_center(ws, row, 3, pn['other_income'])
    _set_num(ws, row, 4, _sum_list(classified.get('other_income_cy', [])))
    _set_num(ws, row, 6, _sum_list(classified.get('other_income_py', [])))
    oi_row = row
    row += 1

    _set_text(ws, row, 1, 'III.', bold=True)
    _set_text(ws, row, 2, 'Total Income (I + II)', bold=True, indent=1)
    _set_num(ws, row, 4, formula=f'=D{rev_row}+D{oi_row}', bold=True)
    _set_num(ws, row, 6, formula=f'=F{rev_row}+F{oi_row}', bold=True)
    _apply_border_row(ws, row, max_col, THIN_TOP_BOTTOM)
    total_income_row = row

    row += 2
    _set_text(ws, row, 1, 'IV.', bold=True)
    _set_text(ws, row, 2, 'EXPENSES', bold=True)
    row += 1

    expense_rows = []

    if 'cost_of_materials' in pn:
        _set_text(ws, row, 2, 'Purchases of Stock-in-Trade', indent=1)
        _set_center(ws, row, 3, pn['cost_of_materials'])
        _set_num(ws, row, 4, _sum_list(classified.get('purchases_cy', [])))
        _set_num(ws, row, 6, _sum_list(classified.get('purchases_py', [])))
        expense_rows.append(row)
        row += 1

    if 'changes_in_inventories' in pn:
        _set_text(ws, row, 2, 'Changes in Inventories of Finished Goods and Stock-in-Trade', indent=1)
        _set_center(ws, row, 3, pn['changes_in_inventories'])
        os_cy = classified.get('opening_stock_cy', 0)
        cs_cy = classified.get('closing_stock_cy', 0)
        os_py = classified.get('opening_stock_py', 0)
        cs_py = classified.get('closing_stock_py', 0)
        change_cy = os_cy - cs_cy
        change_py = os_py - cs_py
        _set_num(ws, row, 4, change_cy)
        _set_num(ws, row, 6, change_py)
        expense_rows.append(row)
        row += 1

    if 'employee_benefits' in pn:
        _set_text(ws, row, 2, 'Employee Benefits Expense', indent=1)
        _set_center(ws, row, 3, pn['employee_benefits'])
        _set_num(ws, row, 4, _sum_list(classified.get('employee_benefits_cy', [])))
        _set_num(ws, row, 6, _sum_list(classified.get('employee_benefits_py', [])))
        expense_rows.append(row)
        row += 1

    if 'finance_costs' in pn:
        _set_text(ws, row, 2, 'Finance Costs', indent=1)
        _set_center(ws, row, 3, pn['finance_costs'])
        _set_num(ws, row, 4, _sum_list(classified.get('finance_costs_cy', [])))
        _set_num(ws, row, 6, _sum_list(classified.get('finance_costs_py', [])))
        expense_rows.append(row)
        row += 1

    if 'depreciation_note' in pn:
        _set_text(ws, row, 2, 'Depreciation and Amortisation Expense', indent=1)
        _set_center(ws, row, 3, pn['depreciation_note'])
        _set_num(ws, row, 4, classified.get('depreciation_cy', 0))
        _set_num(ws, row, 6, classified.get('depreciation_py', 0))
        expense_rows.append(row)
        row += 1

    if 'other_expenses_note' in pn:
        _set_text(ws, row, 2, 'Other Expenses', indent=1)
        _set_center(ws, row, 3, pn['other_expenses_note'])
        _set_num(ws, row, 4, _sum_list(classified.get('other_expenses_cy', [])))
        _set_num(ws, row, 6, _sum_list(classified.get('other_expenses_py', [])))
        expense_rows.append(row)
        row += 1

    row += 1
    _set_text(ws, row, 1, 'V.', bold=True)
    _set_text(ws, row, 2, 'Total Expenses (IV)', bold=True, indent=1)
    if expense_rows:
        exp_formula_cy = '=' + '+'.join([f'D{r}' for r in expense_rows])
        exp_formula_py = '=' + '+'.join([f'F{r}' for r in expense_rows])
    else:
        exp_formula_cy = '=0'
        exp_formula_py = '=0'
    _set_num(ws, row, 4, formula=exp_formula_cy, bold=True)
    _set_num(ws, row, 6, formula=exp_formula_py, bold=True)
    _apply_border_row(ws, row, max_col, THIN_TOP_BOTTOM)
    total_exp_row = row

    row += 2
    _set_text(ws, row, 1, 'VI.', bold=True)
    _set_text(ws, row, 2, 'Profit / (Loss) Before Exceptional Items and Tax (III - V)', bold=True, indent=1)
    _set_num(ws, row, 4, formula=f'=D{total_income_row}-D{total_exp_row}', bold=True)
    _set_num(ws, row, 6, formula=f'=F{total_income_row}-F{total_exp_row}', bold=True)
    pbt_row = row

    row += 1
    _set_text(ws, row, 1, 'VII.', bold=True)
    _set_text(ws, row, 2, 'Exceptional Items', indent=1)
    _set_num(ws, row, 4, 0)
    _set_num(ws, row, 6, 0)
    exc_row = row

    row += 1
    _set_text(ws, row, 1, 'VIII.', bold=True)
    _set_text(ws, row, 2, 'Profit / (Loss) Before Tax (VI - VII)', bold=True, indent=1)
    _set_num(ws, row, 4, formula=f'=D{pbt_row}-D{exc_row}', bold=True)
    _set_num(ws, row, 6, formula=f'=F{pbt_row}-F{exc_row}', bold=True)
    pbt2_row = row

    row += 1
    _set_text(ws, row, 1, 'IX.', bold=True)
    _set_text(ws, row, 2, 'Tax Expense:', indent=1)
    row += 1
    _set_text(ws, row, 2, '(a) Current Tax', indent=2)
    _set_num(ws, row, 4, 0)
    _set_num(ws, row, 6, 0)
    tax_row = row
    row += 1
    _set_text(ws, row, 2, '(b) Deferred Tax', indent=2)
    _set_num(ws, row, 4, 0)
    _set_num(ws, row, 6, 0)
    dt_row = row

    row += 1
    _set_text(ws, row, 1, 'X.', bold=True)
    _set_text(ws, row, 2, 'Profit / (Loss) for the Period (VIII - IX)', bold=True, indent=1)
    _set_num(ws, row, 4, formula=f'=D{pbt2_row}-D{tax_row}-D{dt_row}', bold=True)
    _set_num(ws, row, 6, formula=f'=F{pbt2_row}-F{tax_row}-F{dt_row}', bold=True)
    _apply_border_row(ws, row, max_col, THIN_TOP_DOUBLE_BOTTOM)
    np_row = row

    row_refs['pl_net_profit_cy'] = (ws.title, np_row, 4)
    row_refs['pl_net_profit_py'] = (ws.title, np_row, 6)

    row += 2
    last_note = note_info['last_note_number']
    _set_text(ws, row, 2, f'The accompanying Notes 1 to {last_note} form an integral part of these Financial Statements.',
              italic=True)


def _build_notes(note_sheets, classified, note_info, row_refs, group_ranges):
    for group_name, notes in group_ranges:
        ws = note_sheets.get(group_name)
        if not ws:
            continue

        max_col = 6
        _title_block(ws, classified['entity_name'], group_name, max_col)

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 4
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 2
        ws.column_dimensions['F'].width = 18

        row = 5

        for note in notes:
            row = _write_note(ws, row, note, classified, note_info, row_refs)
            row += 2


def _write_note(ws, row, note, classified, note_info, row_refs):
    note_num = note['number']
    note_id = note['id']
    title = note['title']
    ye_cy = classified.get('year_ending_cy', '')
    ye_py = classified.get('year_ending_py', '')

    _set_text(ws, row, 1, f'Note {note_num}', bold=True)
    _set_text(ws, row, 2, title, bold=True)
    row += 1

    if note_id == 'brief':
        row = _write_brief(ws, row, classified)

    elif note_id == 'policies':
        row = _write_policies(ws, row, classified)

    elif note_id == 'capital':
        row = _write_capital(ws, row, classified, note_info)

    elif note_id in ('lt_borrowings',):
        _set_text(ws, row, 2, 'Particulars', bold=True)
        _set_text(ws, row, 4, f'As at {ye_cy}', bold=True)
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        _set_text(ws, row, 6, f'As at {ye_py}', bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        _apply_border_row(ws, row, 6, THIN_TOP_BOTTOM)
        row += 1
        _set_text(ws, row, 2, 'Secured Loans:', bold=True, indent=1)
        row += 1
        items_cy = classified.get('bank_loan_cy', [])
        items_py = classified.get('bank_loan_py', [])
        row = _write_item_pairs(ws, row, items_cy, items_py)
        _apply_border_row(ws, row - 1, 6, THIN_TOP_DOUBLE_BOTTOM)

    elif note_id == 'st_borrowings':
        _set_text(ws, row, 2, 'Particulars', bold=True)
        _set_text(ws, row, 4, f'As at {ye_cy}', bold=True)
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        _set_text(ws, row, 6, f'As at {ye_py}', bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        _apply_border_row(ws, row, 6, THIN_TOP_BOTTOM)
        row += 1
        all_cy = classified.get('bank_od_cc_cy', []) + classified.get('unsecured_loan_cy', [])
        all_py = classified.get('bank_od_cc_py', []) + classified.get('unsecured_loan_py', [])
        row = _write_item_pairs(ws, row, all_cy, all_py)
        _apply_border_row(ws, row - 1, 6, THIN_TOP_DOUBLE_BOTTOM)

    elif note_id == 'trade_payables':
        _set_text(ws, row, 2, 'Particulars', bold=True)
        _set_text(ws, row, 4, f'As at {ye_cy}', bold=True)
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        _set_text(ws, row, 6, f'As at {ye_py}', bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        _apply_border_row(ws, row, 6, THIN_TOP_BOTTOM)
        row += 1
        _set_text(ws, row, 2, '(i) Due to Micro and Small Enterprises', indent=1)
        _set_num(ws, row, 4, 0)
        _set_num(ws, row, 6, 0)
        row += 1
        _set_text(ws, row, 2, '(ii) Due to Others', indent=1)
        _set_num(ws, row, 4, _sum_list(classified.get('trade_payables_cy', [])))
        _set_num(ws, row, 6, _sum_list(classified.get('trade_payables_py', [])))
        tp_others_row = row
        row += 1
        _set_text(ws, row, 2, 'Total', bold=True, indent=1)
        _set_num(ws, row, 4, formula=f'=D{tp_others_row-1}+D{tp_others_row}', bold=True)
        _set_num(ws, row, 6, formula=f'=F{tp_others_row-1}+F{tp_others_row}', bold=True)
        _apply_border_row(ws, row, 6, THIN_TOP_DOUBLE_BOTTOM)
        row += 1

    elif note_id == 'other_cl':
        _set_text(ws, row, 2, 'Particulars', bold=True)
        _set_text(ws, row, 4, f'As at {ye_cy}', bold=True)
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        _set_text(ws, row, 6, f'As at {ye_py}', bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        _apply_border_row(ws, row, 6, THIN_TOP_BOTTOM)
        row += 1
        all_cy = (classified.get('gst_payable_cy', []) +
                  classified.get('tds_payable_cy', []) +
                  classified.get('other_statutory_cy', []))
        all_py = (classified.get('gst_payable_py', []) +
                  classified.get('tds_payable_py', []) +
                  classified.get('other_statutory_py', []))
        row = _write_item_pairs(ws, row, all_cy, all_py)
        _apply_border_row(ws, row - 1, 6, THIN_TOP_DOUBLE_BOTTOM)

    elif note_id == 'st_provisions':
        _set_text(ws, row, 2, 'Particulars', bold=True)
        _set_text(ws, row, 4, f'As at {ye_cy}', bold=True)
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        _set_text(ws, row, 6, f'As at {ye_py}', bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        _apply_border_row(ws, row, 6, THIN_TOP_BOTTOM)
        row += 1
        row = _write_item_pairs(ws, row,
                                classified.get('provision_tax_cy', []),
                                classified.get('provision_tax_py', []))
        _apply_border_row(ws, row - 1, 6, THIN_TOP_DOUBLE_BOTTOM)

    elif note_id == 'ppe':
        row = _write_ppe_schedule(ws, row, classified)

    elif note_id == 'non_current_investments':
        _set_text(ws, row, 2, 'Particulars', bold=True)
        _set_text(ws, row, 4, f'As at {ye_cy}', bold=True)
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        _set_text(ws, row, 6, f'As at {ye_py}', bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        _apply_border_row(ws, row, 6, THIN_TOP_BOTTOM)
        row += 1
        _set_text(ws, row, 2, 'Other Investments (valued at cost):', bold=True, indent=1)
        row += 1
        row = _write_item_pairs(ws, row,
                                classified.get('gold_investments_cy', []),
                                classified.get('gold_investments_py', []))
        _apply_border_row(ws, row - 1, 6, THIN_TOP_DOUBLE_BOTTOM)

    elif note_id == 'lt_loans_advances':
        _set_text(ws, row, 2, 'Particulars', bold=True)
        _set_text(ws, row, 4, f'As at {ye_cy}', bold=True)
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        _set_text(ws, row, 6, f'As at {ye_py}', bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        _apply_border_row(ws, row, 6, THIN_TOP_BOTTOM)
        row += 1
        _set_text(ws, row, 2, 'Security Deposits (Unsecured, considered good):', bold=True, indent=1)
        row += 1
        row = _write_item_pairs(ws, row,
                                classified.get('security_deposit_cy', []),
                                classified.get('security_deposit_py', []))
        _apply_border_row(ws, row - 1, 6, THIN_TOP_DOUBLE_BOTTOM)

    elif note_id == 'inventories':
        _set_text(ws, row, 2, 'Particulars', bold=True)
        _set_text(ws, row, 4, f'As at {ye_cy}', bold=True)
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        _set_text(ws, row, 6, f'As at {ye_py}', bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        _apply_border_row(ws, row, 6, THIN_TOP_BOTTOM)
        row += 1
        _set_text(ws, row, 2, 'Stock-in-Trade', indent=1)
        inv_cy = classified.get('inventories_cy', 0) or classified.get('closing_stock_cy', 0)
        inv_py = classified.get('inventories_py', 0) or classified.get('closing_stock_py', 0)
        _set_num(ws, row, 4, inv_cy)
        _set_num(ws, row, 6, inv_py)
        row += 1
        _set_text(ws, row, 2, 'Total', bold=True, indent=1)
        _set_num(ws, row, 4, formula=f'=D{row-1}', bold=True)
        _set_num(ws, row, 6, formula=f'=F{row-1}', bold=True)
        _apply_border_row(ws, row, 6, THIN_TOP_DOUBLE_BOTTOM)
        row += 1

    elif note_id == 'trade_receivables':
        _set_text(ws, row, 2, 'Particulars', bold=True)
        _set_text(ws, row, 4, f'As at {ye_cy}', bold=True)
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        _set_text(ws, row, 6, f'As at {ye_py}', bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        _apply_border_row(ws, row, 6, THIN_TOP_BOTTOM)
        row += 1
        _set_text(ws, row, 2, 'Unsecured, considered good:', bold=True, indent=1)
        row += 1
        _set_text(ws, row, 2, 'Outstanding for more than six months', indent=2)
        _set_num(ws, row, 4, 0)
        _set_num(ws, row, 6, 0)
        row += 1
        _set_text(ws, row, 2, 'Others (less than six months)', indent=2)
        _set_num(ws, row, 4, _sum_list(classified.get('trade_receivables_cy', [])))
        _set_num(ws, row, 6, _sum_list(classified.get('trade_receivables_py', [])))
        row += 1
        _set_text(ws, row, 2, 'Total', bold=True, indent=1)
        _set_num(ws, row, 4, formula=f'=D{row-2}+D{row-1}', bold=True)
        _set_num(ws, row, 6, formula=f'=F{row-2}+F{row-1}', bold=True)
        _apply_border_row(ws, row, 6, THIN_TOP_DOUBLE_BOTTOM)
        row += 1

    elif note_id == 'cash_bank':
        _set_text(ws, row, 2, 'Particulars', bold=True)
        _set_text(ws, row, 4, f'As at {ye_cy}', bold=True)
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        _set_text(ws, row, 6, f'As at {ye_py}', bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        _apply_border_row(ws, row, 6, THIN_TOP_BOTTOM)
        row += 1

        start_row = row
        _set_text(ws, row, 2, '(a) Cash on Hand', bold=True, indent=1)
        row += 1
        for item in classified.get('cash_hand_cy', []):
            _set_text(ws, row, 2, item['name'], indent=2)
            _set_num(ws, row, 4, item['amount'])
            py_match = _find_py_match(item['name'], classified.get('cash_hand_py', []))
            _set_num(ws, row, 6, py_match)
            row += 1
        if not classified.get('cash_hand_cy'):
            _set_text(ws, row, 2, 'Cash on Hand', indent=2)
            _set_num(ws, row, 4, 0)
            _set_num(ws, row, 6, _sum_list(classified.get('cash_hand_py', [])))
            row += 1

        _set_text(ws, row, 2, '(b) Balances with Banks', bold=True, indent=1)
        row += 1
        for item in classified.get('bank_balance_cy', []):
            _set_text(ws, row, 2, item['name'], indent=2)
            _set_num(ws, row, 4, item['amount'])
            py_match = _find_py_match(item['name'], classified.get('bank_balance_py', []))
            _set_num(ws, row, 6, py_match)
            row += 1
        if not classified.get('bank_balance_cy') and classified.get('bank_balance_py'):
            for item in classified.get('bank_balance_py', []):
                _set_text(ws, row, 2, item['name'], indent=2)
                _set_num(ws, row, 4, 0)
                _set_num(ws, row, 6, item['amount'])
                row += 1

        if classified.get('fixed_deposit_cy') or classified.get('fixed_deposit_py'):
            _set_text(ws, row, 2, '(c) Other Bank Balances', bold=True, indent=1)
            row += 1
            _set_text(ws, row, 2, 'Fixed Deposits with Banks', indent=2)
            _set_num(ws, row, 4, _sum_list(classified.get('fixed_deposit_cy', [])))
            _set_num(ws, row, 6, _sum_list(classified.get('fixed_deposit_py', [])))
            row += 1

        _set_text(ws, row, 2, 'Total', bold=True, indent=1)
        _set_num(ws, row, 4, formula=f'=SUM(D{start_row}:D{row-1})', bold=True)
        _set_num(ws, row, 6, formula=f'=SUM(F{start_row}:F{row-1})', bold=True)
        _apply_border_row(ws, row, 6, THIN_TOP_DOUBLE_BOTTOM)
        row += 1

    elif note_id == 'st_loans_advances':
        _set_text(ws, row, 2, 'Particulars', bold=True)
        _set_text(ws, row, 4, f'As at {ye_cy}', bold=True)
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        _set_text(ws, row, 6, f'As at {ye_py}', bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        _apply_border_row(ws, row, 6, THIN_TOP_BOTTOM)
        row += 1
        _set_text(ws, row, 2, 'Balances with statutory / government authorities:', bold=True, indent=1)
        row += 1
        all_cy = classified.get('gst_input_cy', []) + classified.get('tds_advance_tax_cy', [])
        all_py = classified.get('gst_input_py', []) + classified.get('tds_advance_tax_py', [])
        row = _write_item_pairs(ws, row, all_cy, all_py)
        _apply_border_row(ws, row - 1, 6, THIN_TOP_DOUBLE_BOTTOM)

    elif note_id == 'other_ca':
        _set_text(ws, row, 2, 'Particulars', bold=True)
        _set_text(ws, row, 4, f'As at {ye_cy}', bold=True)
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        _set_text(ws, row, 6, f'As at {ye_py}', bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        _apply_border_row(ws, row, 6, THIN_TOP_BOTTOM)
        row += 1
        row = _write_item_pairs(ws, row,
                                classified.get('other_ca_cy', []),
                                classified.get('other_ca_py', []))
        _apply_border_row(ws, row - 1, 6, THIN_TOP_DOUBLE_BOTTOM)

    elif note_id in ('revenue', 'other_income', 'cost_of_materials',
                      'employee_benefits', 'finance_costs', 'other_expenses_note'):
        ye_label_cy = f'For the year ended {ye_cy}'
        ye_label_py = f'For the year ended {ye_py}'
        _set_text(ws, row, 2, 'Particulars', bold=True)
        _set_text(ws, row, 4, ye_label_cy, bold=True)
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        _set_text(ws, row, 6, ye_label_py, bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        _apply_border_row(ws, row, 6, THIN_TOP_BOTTOM)
        row += 1

        key_map = {
            'revenue': ('revenue_from_operations', 'revenue_from_operations'),
            'other_income': ('other_income', 'other_income'),
            'cost_of_materials': ('purchases', 'purchases'),
            'employee_benefits': ('employee_benefits', 'employee_benefits'),
            'finance_costs': ('finance_costs', 'finance_costs'),
            'other_expenses_note': ('other_expenses', 'other_expenses'),
        }
        cy_key, py_key = key_map[note_id]
        row = _write_item_pairs(ws, row,
                                classified.get(f'{cy_key}_cy', []),
                                classified.get(f'{py_key}_py', []))
        _apply_border_row(ws, row - 1, 6, THIN_TOP_DOUBLE_BOTTOM)

    elif note_id == 'changes_in_inventories':
        ye_label_cy = f'For the year ended {ye_cy}'
        ye_label_py = f'For the year ended {ye_py}'
        _set_text(ws, row, 2, 'Particulars', bold=True)
        _set_text(ws, row, 4, ye_label_cy, bold=True)
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        _set_text(ws, row, 6, ye_label_py, bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        _apply_border_row(ws, row, 6, THIN_TOP_BOTTOM)
        row += 1

        os_cy = classified.get('opening_stock_cy', 0)
        cs_cy = classified.get('closing_stock_cy', 0)
        os_py = classified.get('opening_stock_py', 0)
        cs_py = classified.get('closing_stock_py', 0)

        _set_text(ws, row, 2, 'Opening Stock', indent=1)
        _set_num(ws, row, 4, os_cy)
        _set_num(ws, row, 6, os_py)
        row += 1
        _set_text(ws, row, 2, 'Less: Closing Stock', indent=1)
        _set_num(ws, row, 4, cs_cy)
        _set_num(ws, row, 6, cs_py)
        row += 1
        _set_text(ws, row, 2, 'Total', bold=True, indent=1)
        _set_num(ws, row, 4, formula=f'=D{row-2}-D{row-1}', bold=True)
        _set_num(ws, row, 6, formula=f'=F{row-2}-F{row-1}', bold=True)
        _apply_border_row(ws, row, 6, THIN_TOP_DOUBLE_BOTTOM)
        row += 1

    elif note_id == 'depreciation_note':
        ye_label_cy = f'For the year ended {ye_cy}'
        ye_label_py = f'For the year ended {ye_py}'
        _set_text(ws, row, 2, 'Particulars', bold=True)
        _set_text(ws, row, 4, ye_label_cy, bold=True)
        ws.cell(row, 4).alignment = Alignment(horizontal='center')
        _set_text(ws, row, 6, ye_label_py, bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        _apply_border_row(ws, row, 6, THIN_TOP_BOTTOM)
        row += 1
        dep_cy = classified.get('depreciation_cy', 0)
        dep_py = classified.get('depreciation_py', 0)
        if dep_cy or dep_py:
            _set_text(ws, row, 2, 'Depreciation on Property, Plant and Equipment', indent=1)
            _set_num(ws, row, 4, dep_cy)
            _set_num(ws, row, 6, dep_py)
            row += 1
            _set_text(ws, row, 2, 'Total', bold=True, indent=1)
            _set_num(ws, row, 4, formula=f'=D{row-1}', bold=True)
            _set_num(ws, row, 6, formula=f'=F{row-1}', bold=True)
            _apply_border_row(ws, row, 6, THIN_TOP_DOUBLE_BOTTOM)
        else:
            _set_text(ws, row, 2, 'Depreciation has not been provided in the books of account.', italic=True)
        row += 1

    elif note_id == 'previous_year':
        _set_text(ws, row, 2,
                  'Previous year figures have been regrouped / reclassified wherever necessary '
                  'to confirm to the current year classification / presentation. Such regrouping '
                  'does not affect the previously reported net profit or equity of the entity.')
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[row].height = 45
        row += 1

    elif note_id == 'rounding_off':
        _set_text(ws, row, 2,
                  'All amounts in the financial statements are presented in Indian Rupees (Rs.) '
                  'and rounded off to the nearest rupee unless otherwise stated.')
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[row].height = 30
        row += 1

        row += 1
        last_note = note_info['last_note_number']
        _set_text(ws, row, 2,
                  f'Signatures to Notes 1 to {last_note}', bold=True)
        row += 2

        _set_text(ws, row, 1, 'For and on behalf of', bold=True)
        _set_text(ws, row, 4, 'For and on behalf of the Entity', bold=True)
        row += 1
        _set_text(ws, row, 1, '____________________')
        _set_text(ws, row, 4, '____________________')
        row += 1
        _set_text(ws, row, 1, 'Chartered Accountants')
        prop_name = classified.get('proprietor_name', '—')
        _set_text(ws, row, 4, f'Shri/Smt. {prop_name}')
        row += 1
        _set_text(ws, row, 1, 'Firm Registration No.: —')
        _set_text(ws, row, 4, '(Proprietor)')
        row += 2
        _set_text(ws, row, 1, '____________________')
        _set_text(ws, row, 4, '')
        row += 1
        _set_text(ws, row, 1, '(Name of Partner)')
        row += 1
        _set_text(ws, row, 1, 'Partner')
        row += 1
        _set_text(ws, row, 1, 'Membership No.: —')
        row += 2
        _set_text(ws, row, 1, 'UDIN: —')
        _set_text(ws, row, 4, f'PAN: —')
        row += 1
        _set_text(ws, row, 1, 'Place: —')
        _set_text(ws, row, 4, 'Place: —')
        row += 1
        _set_text(ws, row, 1, 'Date: —')
        _set_text(ws, row, 4, 'Date: —')

    return row


def _write_brief(ws, row, classified):
    entity = classified['entity_name']
    constitution = classified.get('constitution', 'proprietorship')
    prop_name = classified.get('proprietor_name', '—')

    items = [
        ('1.1', 'Name of the Entity', entity),
        ('1.2', 'Constitution', constitution.title() + ' Concern'),
        ('1.3', 'Name of Proprietor' if constitution == 'proprietorship' else 'Names of Partners',
         f'Shri/Smt. {prop_name}' if prop_name else '—'),
        ('1.4', 'Nature of Business', '— (to be specified)'),
        ('1.5', 'Address of Registered Office', '—'),
        ('1.6', 'PAN', '—'),
        ('1.7', 'GSTIN', '—'),
        ('1.8', 'Date of Commencement of Business', '—'),
        ('1.9', 'NCE Level', 'Level IV'),
    ]

    for sr, label, value in items:
        _set_text(ws, row, 1, sr)
        _set_text(ws, row, 2, label)
        _set_text(ws, row, 4, value)
        row += 1

    return row


def _write_policies(ws, row, classified):
    constitution = classified.get('constitution', 'proprietorship')
    has_ppe = bool(classified.get('ppe_cy') or classified.get('ppe_py'))
    has_dep = classified.get('depreciation_cy', 0) > 0 or classified.get('depreciation_py', 0) > 0
    has_inv = classified.get('inventories_cy', 0) > 0 or classified.get('closing_stock_cy', 0) > 0

    _set_text(ws, row, 2, '2.1 Basis of Preparation', bold=True)
    row += 1
    _set_text(ws, row, 2,
              'The financial statements have been prepared in accordance with the Accounting Standards '
              'as prescribed under Section 133 of the Companies Act, 2013 (to the extent applicable to '
              'Non-Corporate Entities) and the Guidance Note on Accounting Standards for NCEs issued by '
              'ICAI. The financial statements have been prepared under the historical cost convention on '
              'accrual basis of accounting.')
    ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 60
    row += 2

    _set_text(ws, row, 2, '2.2 Use of Estimates', bold=True)
    row += 1
    _set_text(ws, row, 2,
              'The preparation of financial statements requires management to make estimates and '
              'assumptions that affect the reported amounts of assets and liabilities and disclosure of '
              'contingent liabilities at the date of the financial statements and the reported amounts of '
              'revenues and expenses during the reporting period.')
    ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 55
    row += 2

    if has_ppe:
        _set_text(ws, row, 2, '2.3 Property, Plant and Equipment', bold=True)
        row += 1
        if has_dep:
            txt = ('Property, Plant and Equipment are stated at cost less accumulated depreciation. '
                   'Depreciation is provided on Written Down Value / Straight Line Method basis at the '
                   'rates prescribed under the Income Tax Act, 1961.')
        else:
            txt = ('Property, Plant and Equipment are stated at cost. Depreciation has not been provided '
                   'in the books of account during the year. The proprietor intends to provide depreciation '
                   'on useful life basis as prescribed under Schedule II to the Companies Act, 2013 in the '
                   'ensuing year. The impact on the financial statements on account of non-provision of '
                   'depreciation has not been ascertained.')
        _set_text(ws, row, 2, txt)
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[row].height = 60
        row += 2

    if has_inv:
        _set_text(ws, row, 2, '2.4 Inventories', bold=True)
        row += 1
        _set_text(ws, row, 2,
                  'Inventories are valued at the lower of cost and net realisable value. Cost is '
                  'determined on First-In-First-Out (FIFO) basis.')
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[row].height = 30
        row += 2

    _set_text(ws, row, 2, '2.5 Revenue Recognition', bold=True)
    row += 1
    _set_text(ws, row, 2,
              'Revenue is recognised when the significant risks and rewards of ownership of goods have '
              'passed to the buyer. Revenue from services is recognised proportionately with the degree '
              'of completion of services.')
    ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 45
    row += 2

    if constitution == 'proprietorship':
        _set_text(ws, row, 2, '2.6 Income Tax', bold=True)
        row += 1
        _set_text(ws, row, 2,
                  'The entity being a proprietorship concern, the income of the entity is assessable in '
                  'the personal hands of the proprietor under the Income Tax Act, 1961. Accordingly, no '
                  'provision for income tax or deferred tax has been made in these financial statements.')
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[row].height = 45
        row += 2

    _set_text(ws, row, 2, '2.7 Cash Flow Statement', bold=True)
    row += 1
    _set_text(ws, row, 2,
              'The entity being a Level IV Non-Corporate Entity, Cash Flow Statement is not prepared '
              'in accordance with the exemption available under AS 3.')
    ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 30
    row += 2

    return row


def _write_capital(ws, row, classified, note_info):
    constitution = classified.get('constitution', 'proprietorship')
    entity = classified['entity_name']
    ye_cy = classified.get('year_ending_cy', '')
    ye_py = classified.get('year_ending_py', '')
    prop_name = classified.get('proprietor_name', '—')

    cap_cy = classified.get('capital_cy', {})
    cap_py = classified.get('capital_py', {})

    if constitution == 'proprietorship':
        _set_text(ws, row, 2, f"3(a) Proprietor's Capital Account", bold=True)
        row += 1

        headers = ['Sr.', 'Name', 'Share %', 'Opening Balance',
                   'Capital Introduced', 'Net Profit/(Loss)',
                   'Interest on Capital', 'Withdrawals', 'Closing Balance']

        for c, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(c)].width = max(14, len(h) + 2)
            _set_text(ws, row, c, h, bold=True)
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 25
        _apply_border_row(ws, row, len(headers), THIN_TOP_BOTTOM)
        row += 1

        ye_label = f'For the year ended {ye_cy}'
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        _set_text(ws, row, 1, ye_label, bold=True, italic=True)
        row += 1

        _set_center(ws, row, 1, 1)
        _set_text(ws, row, 2, f'Shri/Smt. {prop_name}')
        _set_text(ws, row, 3, '100%')
        _set_num(ws, row, 4, cap_cy.get('opening', 0))
        _set_num(ws, row, 5, cap_cy.get('capital_introduced', 0))
        _set_num(ws, row, 6, cap_cy.get('net_profit', 0) or classified.get('net_profit_cy', 0))
        _set_num(ws, row, 7, cap_cy.get('interest_on_capital', 0))
        _set_num(ws, row, 8, cap_cy.get('drawings', 0))
        _set_num(ws, row, 9,
                 formula=f'=D{row}+E{row}+F{row}+G{row}-H{row}', bold=True)
        _apply_border_row(ws, row, len(headers), THIN_BOTTOM)
        row += 1

        if cap_py.get('opening', 0) or cap_py.get('closing', 0):
            ye_label_py = f'For the year ended {ye_py}'
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            _set_text(ws, row, 1, ye_label_py, bold=True, italic=True)
            row += 1

            _set_center(ws, row, 1, 1)
            _set_text(ws, row, 2, f'Shri/Smt. {prop_name}')
            _set_text(ws, row, 3, '100%')
            _set_num(ws, row, 4, cap_py.get('opening', 0))
            _set_num(ws, row, 5, cap_py.get('capital_introduced', 0))
            _set_num(ws, row, 6, cap_py.get('net_profit', 0) or classified.get('net_profit_py', 0))
            _set_num(ws, row, 7, cap_py.get('interest_on_capital', 0))
            _set_num(ws, row, 8, cap_py.get('drawings', 0))
            _set_num(ws, row, 9,
                     formula=f'=D{row}+E{row}+F{row}+G{row}-H{row}', bold=True)
            _apply_border_row(ws, row, len(headers), THIN_BOTTOM)
            row += 1

        row += 1
        _set_text(ws, row, 2,
                  '3(b) Not applicable — Entity being proprietorship, separate Current Account not maintained.',
                  italic=True)
        row += 1

    return row


def _write_ppe_schedule(ws, row, classified):
    ye_cy = classified.get('year_ending_cy', '')
    ye_py = classified.get('year_ending_py', '')
    has_dep = classified.get('depreciation_cy', 0) > 0

    ppe_items_cy = classified.get('ppe_cy', [])
    ppe_items_py = classified.get('ppe_py', [])

    all_names = list({i['name'] for i in ppe_items_cy + ppe_items_py})
    all_names.sort()

    if not has_dep:
        headers = ['Sr.', 'Asset', f'Opening GB\n1 April', 'Additions', f'Closing GB\n31 March',
                   f'Net Block\n{ye_cy}', f'Net Block\n{ye_py}']
        col_count = len(headers)

        for c, h in enumerate(headers, 1):
            _set_text(ws, row, c, h, bold=True)
            ws.cell(row, c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 28
        for c in range(3, col_count + 1):
            ws.column_dimensions[get_column_letter(c)].width = 15
        _apply_border_row(ws, row, col_count, THIN_TOP_BOTTOM)
        ws.row_dimensions[row].height = 30
        row += 1

        sr = 1
        data_start = row
        for name in all_names:
            cy_item = next((i for i in ppe_items_cy if i['name'] == name), None)
            py_item = next((i for i in ppe_items_py if i['name'] == name), None)

            cy_val = cy_item['amount'] if cy_item else 0
            py_val = py_item['amount'] if py_item else 0
            addition = cy_val - py_val if cy_val > py_val else 0

            _set_center(ws, row, 1, sr)
            _set_text(ws, row, 2, name)
            _set_num(ws, row, 3, py_val)
            _set_num(ws, row, 4, addition)
            _set_num(ws, row, 5, formula=f'=C{row}+D{row}')
            _set_num(ws, row, 6, formula=f'=E{row}')
            _set_num(ws, row, 7, py_val)
            sr += 1
            row += 1

        data_end = row - 1
        _set_text(ws, row, 2, 'Total', bold=True)
        for c in range(3, col_count + 1):
            _set_num(ws, row, c,
                     formula=f'=SUM({get_column_letter(c)}{data_start}:{get_column_letter(c)}{data_end})',
                     bold=True)
        _apply_border_row(ws, row, col_count, THIN_TOP_DOUBLE_BOTTOM)
        row += 1

        row += 1
        _set_text(ws, row, 2,
                  'Note: Depreciation has not been provided in the books of account. The amounts shown '
                  'represent Gross Block (original cost). The proprietor intends to provide depreciation '
                  'on useful life basis in the ensuing year. Impact not ascertained.',
                  italic=True)
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[row].height = 45
        row += 1

    else:
        headers = ['Sr.', 'Asset',
                   f'Opening GB\n1 April', 'Additions', f'Closing GB\n31 March',
                   f'Opening Dep\n1 April', 'Dep for Year', f'Closing Dep\n31 March',
                   f'Net Block\n{ye_cy}', f'Net Block\n{ye_py}']
        col_count = len(headers)

        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        _set_text(ws, row, 3, 'GROSS BLOCK', bold=True)
        ws.cell(row, 3).alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
        _set_text(ws, row, 6, 'ACCUMULATED DEPRECIATION', bold=True)
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=10)
        _set_text(ws, row, 9, 'NET BLOCK', bold=True)
        ws.cell(row, 9).alignment = Alignment(horizontal='center')
        _apply_border_row(ws, row, col_count, THIN_TOP_BOTTOM)
        row += 1

        for c, h in enumerate(headers, 1):
            _set_text(ws, row, c, h, bold=True)
            ws.cell(row, c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 28
        for c in range(3, col_count + 1):
            ws.column_dimensions[get_column_letter(c)].width = 14
        _apply_border_row(ws, row, col_count, THIN_BOTTOM)
        ws.row_dimensions[row].height = 30
        row += 1

        sr = 1
        data_start = row
        for name in all_names:
            cy_item = next((i for i in ppe_items_cy if i['name'] == name), None)
            py_item = next((i for i in ppe_items_py if i['name'] == name), None)
            cy_val = cy_item['amount'] if cy_item else 0
            py_val = py_item['amount'] if py_item else 0

            _set_center(ws, row, 1, sr)
            _set_text(ws, row, 2, name)
            _set_num(ws, row, 3, py_val)
            _set_num(ws, row, 4, max(cy_val - py_val, 0))
            _set_num(ws, row, 5, formula=f'=C{row}+D{row}')
            _set_num(ws, row, 6, 0)
            _set_num(ws, row, 7, 0)
            _set_num(ws, row, 8, formula=f'=F{row}+G{row}')
            _set_num(ws, row, 9, formula=f'=E{row}-H{row}')
            _set_num(ws, row, 10, formula=f'=C{row}-F{row}')
            sr += 1
            row += 1

        data_end = row - 1
        _set_text(ws, row, 2, 'Total', bold=True)
        for c in range(3, col_count + 1):
            _set_num(ws, row, c,
                     formula=f'=SUM({get_column_letter(c)}{data_start}:{get_column_letter(c)}{data_end})',
                     bold=True)
        _apply_border_row(ws, row, col_count, THIN_TOP_DOUBLE_BOTTOM)
        row += 1

    row += 1
    _set_text(ws, row, 2,
              'No revaluation of Property, Plant and Equipment has been carried out during the year.',
              italic=True)
    row += 1
    _set_text(ws, row, 2,
              'There are no Intangible Assets, Capital Work-in-Progress, or Intangible Assets under '
              'Development.',
              italic=True)
    row += 1
    _set_text(ws, row, 2,
              'All assets are owned by the entity and are free from any charge or encumbrance unless '
              'otherwise stated.',
              italic=True)
    row += 1

    return row


def _write_item_pairs(ws, row, cy_items, py_items):
    all_names = []
    seen = set()
    for i in cy_items:
        n = i['name']
        if n.lower() not in seen:
            all_names.append(n)
            seen.add(n.lower())
    for i in py_items:
        n = i['name']
        if n.lower() not in seen:
            all_names.append(n)
            seen.add(n.lower())

    start_row = row
    for name in all_names:
        _set_text(ws, row, 2, name, indent=2)
        cy_val = next((i['amount'] for i in cy_items if i['name'].lower() == name.lower()), 0)
        py_val = next((i['amount'] for i in py_items if i['name'].lower() == name.lower()), 0)
        _set_num(ws, row, 4, cy_val)
        _set_num(ws, row, 6, py_val)
        row += 1

    if not all_names:
        _set_text(ws, row, 2, '—', indent=2)
        _set_num(ws, row, 4, 0)
        _set_num(ws, row, 6, 0)
        row += 1
        start_row = row - 1

    _set_text(ws, row, 2, 'Total', bold=True, indent=1)
    if row - 1 == start_row:
        _set_num(ws, row, 4, formula=f'=D{start_row}', bold=True)
        _set_num(ws, row, 6, formula=f'=F{start_row}', bold=True)
    else:
        _set_num(ws, row, 4, formula=f'=SUM(D{start_row}:D{row-1})', bold=True)
        _set_num(ws, row, 6, formula=f'=SUM(F{start_row}:F{row-1})', bold=True)
    row += 1

    return row


def _find_py_match(name, py_items):
    name_lower = name.lower()
    for item in py_items:
        if item['name'].lower() == name_lower:
            return item['amount']
    return 0


def _normalize_fonts(wb):
    from openpyxl.cell.cell import Cell, MergedCell
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.font = Font(
                    name='Calibri Light',
                    size=11,
                    bold=cell.font.bold if cell.font else False,
                    italic=cell.font.italic if cell.font else False,
                    underline=None,
                    color=cell.font.color if cell.font else None,
                )
